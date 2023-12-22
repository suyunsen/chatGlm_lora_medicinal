"""
 -*- coding: utf-8 -*-
time: 2023/11/22 18:43
author: suyunsen
email: suyunsen2023@gmail.com
"""


import openpyxl
import json
import re

import random
all_data = []
def get_all_data():

    global all_data
    # 打开Excel文件
    workbook = openpyxl.load_workbook('./data/train.xlsx')

    # 选择要读取的工作表
    worksheet = workbook['单事项信息']

    # 项目名称[7]，网上办理流程说明[35],窗口办理流程说明[36],地区[1]

    filter_data = ['广东省', '佛山市']

    columns = ['Human', 'Model']
    template_pre = [['我要办', '我要办', '我打算办', '办', '怎么办'], '请问您办理那个区域的', '我在',
                    '请问您是首次办理、到期换证还是补办？', '',
                    '请问您是想要线上办理还是窗口办理?', '', '根据您的需求，您可以有如下方式办理：']
    template_suf = [['，我要办理那些东西', '', '有那些问题要办', '', ''], '，您所在的镇街是那里？', '', '', '', '', '', '']

    # 正则表达

    exchange_r = r"换发"
    comple_r = r"补发"

    # 遍历每一行数据
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        dialogue_data = []
        envent = row[7]

        wet_d = row[35]

        xianxia_d = row[36]

        area = row[1]

        row_data = [''] * len(template_pre)

        for i in range(2):
            row_data[i] = envent

        row_data[2] = area

        if re.search(exchange_r, envent):
            row_data[4] = '换证'
        elif re.search(comple_r, envent):
            row_data[4] = '补办'
        else:
            row_data[4] = '首次'

        row_data[6] = '窗口'
        row_data[7] = xianxia_d if xianxia_d is not None else "无窗口办理流程"
        if xianxia_d is None:
            continue
        # 遍历要读取的列
        flag = False

        dialogue = {}

        cnt = 0

        mid_st = '，这个'
        for i in range(8):
            cnt += 1
            ans = ''
            if i == 0:
                count = random.randint(0, 4)
                ans = template_pre[i][count] + row_data[i]
                if count == 2:
                    ans += mid_st + row_data[i]
                ans += template_suf[i][count]
            else:
                ans = template_pre[i] + row_data[i] + template_suf[i]

            if cnt == 2 and area in filter_data:
                flag = True
                continue
            if cnt == 3 and flag:
                continue
            dialogue[columns[i % 2]] = ans
            if len(dialogue) == 2:
                dialogue_data.append(dialogue)
                dialogue = {}

        all_data.append(dialogue_data)

        dialogue_data = {}

        cnt = 0
        flag = False

        row_data[6] = '线上'
        row_data[7] = wet_d if wet_d is not None else "无线上办理流程"
        dialogue_data = []
        for i in range(8):
            cnt += 1
            ans = ''
            if i == 0:
                count = random.randint(0, 4)
                ans = template_pre[i][count] + row_data[i]
                if count == 2:
                    ans += mid_st + row_data[i]
                ans += template_suf[i][count]
            else:
                ans = template_pre[i] + row_data[i] + template_suf[i]

            if cnt == 2 and area in filter_data:
                flag = True
                continue
            if cnt == 3 and flag:
                continue
            dialogue[columns[i % 2]] = ans
            if len(dialogue) == 2:
                dialogue_data.append(dialogue)
                dialogue = {}

        all_data.append(dialogue_data)

def get_al_data():
    global all_data
    # 打开Excel文件
    workbook = openpyxl.load_workbook('./data/train.xlsx')

    # 选择要读取的工作表
    worksheet = workbook['单事项信息']

    # 项目名称[7]，网上办理流程说明[35],窗口办理流程说明[36],地区[1]

    filter_data = ['广东省', '佛山市']

    columns = ['Human', 'Model']
    template_pre = [['我要办', '我要办', '我打算办', '办', '怎么办'], ['请问您在那个区域，您是线上办理还是窗口办理，您是首次办理、到期换证还是补办？','请问您是线上办理还是窗口办理，您是首次办理、到期换证还是补办？'],'','根据您的需求，您可以有如下方式办理：']
    template_suf = [['，我要办理那些东西', '', '有那些问题要办', '', ''], '','', '']

    # 正则表达

    exchange_r = r"换发"
    comple_r = r"补发"

    # 遍历每一行数据
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        dialogue_data = []
        envent = row[7]

        wet_d = row[35]

        xianxia_d = row[36]

        area = row[1]

        row_data = [''] * 10

        for i in range(2):
            row_data[i] = envent

        row_data[2] = area

        if re.search(exchange_r, envent):
            row_data[4] = '换证'
        elif re.search(comple_r, envent):
            row_data[4] = '补办'
        else:
            row_data[4] = '首次'

        row_data[6] = '窗口'
        row_data[7] = xianxia_d if xianxia_d is not None else "无窗口办理流程"

        mid_st = '，这个'
        if xianxia_d is None:
            continue
        else :

            # 遍历要读取的列
            flag = False

            dialogue = {}

            cnt = 0

            new_ans = ['','']
            if area in filter_data:
                new_ans.append(row_data[6] + "，"+row_data[4])
            else:
                new_ans.append("在"+row_data[2]+"，"+row_data[6] + "，"+row_data[4])
            new_ans.append(xianxia_d)
            for i in range(4):
                cnt += 1
                ans = ''
                if i == 0:
                    count = random.randint(0, 4)
                    ans = template_pre[i][count] + row_data[i]
                    if count == 2:
                        ans += mid_st + row_data[i]
                    ans += template_suf[i][count]
                elif i == 1:
                    if area in filter_data:
                        ans = template_pre[i][1] + template_suf[i]
                    else:
                        ans = template_pre[i][0] + template_suf[i]
                else:
                    ans = template_pre[i] + new_ans[i] + template_suf[i]

                dialogue[columns[i % 2]] = ans
                if len(dialogue) == 2:
                    dialogue_data.append(dialogue)
                    dialogue = {}

            all_data.append(dialogue_data)



        dialogue_data = []
        cnt = 0
        flag = False

        row_data[6] = '线上'
        row_data[7] = wet_d if wet_d is not None else "无线上办理流程"
        if wet_d is None:
            continue
        new_ans = ['', '']
        if area in filter_data:
            new_ans.append(row_data[6] + "，" + row_data[4])
        else:
            new_ans.append("在" + row_data[2] + "，" + row_data[6] + "，" + row_data[4])
        new_ans.append(wet_d)
        for i in range(4):
            cnt += 1
            ans = ''
            if i == 0:
                count = random.randint(0, 4)
                ans = template_pre[i][count] + row_data[i]
                if count == 2:
                    ans += mid_st + row_data[i]
                ans += template_suf[i][count]
            elif i == 1:
                if area in filter_data:
                    ans = template_pre[i][1] + template_suf[i]
                else:
                    ans = template_pre[i][0] + template_suf[i]
            else:
                ans = template_pre[i] + new_ans[i] + template_suf[i]

            dialogue[columns[i % 2]] = ans
            if len(dialogue) == 2:
                dialogue_data.append(dialogue)
                dialogue = {}

        all_data.append(dialogue_data)
def get_area_data():
    global all_data
    # 打开Excel文件
    workbook = openpyxl.load_workbook('./data/train.xlsx')

    # 选择要读取的工作表
    worksheet = workbook['单事项信息']

    # 项目名称[7]，网上办理流程说明[35],窗口办理流程说明[36],地区[1]

    filter_data = ['广东省', '佛山市']

    columns = ['Human', 'Model']
    template_pre = [['我要办', '我要办', '我打算办', '办', '怎么办'],
                     '请问您是线上办理还是窗口办理，您是首次办理、到期换证还是补办？', '',
                    '根据您的需求，您可以有如下方式办理：']
    template_suf = [['，我要办理那些东西', '', '有那些问题要办', '', ''], '', '', '']

    # 正则表达

    exchange_r = r"换发"
    comple_r = r"补发"

    # 遍历每一行数据
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        dialogue_data = []
        envent = row[7]

        wet_d = row[35]

        xianxia_d = row[36]

        area = row[1]

        row_data = [''] * 10

        for i in range(2):
            row_data[i] = envent

        row_data[2] = area

        if re.search(exchange_r, envent):
            row_data[4] = '换证'
        elif re.search(comple_r, envent):
            row_data[4] = '补办'
        else:
            row_data[4] = '首次'

        row_data[6] = '窗口'
        row_data[7] = xianxia_d if xianxia_d is not None else "无窗口办理流程"

        mid_st = '，这个'
        if xianxia_d is None:
            continue
        else:

            # 遍历要读取的列
            flag = False

            dialogue = {}

            cnt = 0

            new_ans = ['', '']

            new_ans.append(row_data[6] + "，" + row_data[4])
            new_ans.append(xianxia_d)

            new_areas = area
            if area in filter_data:
                new_areas = ''
            for i in range(4):
                cnt += 1
                ans = ''
                if i == 0:
                    count = random.randint(0, 4)
                    ans = template_pre[i][count] + new_areas + row_data[i]
                    if count == 2:
                        ans += mid_st + row_data[i]
                    ans += template_suf[i][count]
                else:
                    ans = template_pre[i] + new_ans[i] + template_suf[i]

                dialogue[columns[i % 2]] = ans
                if len(dialogue) == 2:
                    dialogue_data.append(dialogue)
                    dialogue = {}

            all_data.append(dialogue_data)

        dialogue_data = []
        cnt = 0
        flag = False

        row_data[6] = '线上'
        row_data[7] = wet_d if wet_d is not None else "无线上办理流程"
        if wet_d is None:
            continue
        new_ans = ['', '']

        new_ans.append(row_data[6] + "，" + row_data[4])
        new_areas = area
        if area in filter_data:
            new_areas = ''
        new_ans.append(wet_d)
        for i in range(4):
            cnt += 1
            ans = ''
            if i == 0:
                count = random.randint(0, 4)
                ans = template_pre[i][count] + new_areas+ row_data[i]
                if count == 2:
                    ans += mid_st + row_data[i]
                ans += template_suf[i][count]
            elif i == 1:
                if area in filter_data:
                    ans = template_pre[i][1] + template_suf[i]
                else:
                    ans = template_pre[i][0] + template_suf[i]
            else:
                ans = template_pre[i] + new_ans[i] + template_suf[i]

            dialogue[columns[i % 2]] = ans
            if len(dialogue) == 2:
                dialogue_data.append(dialogue)
                dialogue = {}

        all_data.append(dialogue_data)

def get_windowsOrOffilen_data():
    global all_data
    # 打开Excel文件
    workbook = openpyxl.load_workbook('./data/train.xlsx')

    # 选择要读取的工作表
    worksheet = workbook['单事项信息']

    # 项目名称[7]，网上办理流程说明[35],窗口办理流程说明[36],地区[1]

    filter_data = ['广东省', '佛山市']

    columns = ['Human', 'Model']
    template_pre = [['我要', '我要', '我打算', '我想', '怎么'],
                    ['请问您在那个区域，您是首次办理、到期换证还是补办？','请问您是首次办理、到期换证还是补办？'], '',
                    '根据您的需求，您可以有如下方式办理：']
    template_suf = [['，我要办理那些东西', '', '有那些问题要办', '', ''], '', '', '']

    # 正则表达

    exchange_r = r"换发"
    comple_r = r"补发"

    # 遍历每一行数据
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        dialogue_data = []
        envent = row[7]

        wet_d = row[35]

        xianxia_d = row[36]

        area = row[1]

        row_data = [''] * 10

        for i in range(2):
            row_data[i] = envent

        row_data[2] = area

        if re.search(exchange_r, envent):
            row_data[4] = '换证'
        elif re.search(comple_r, envent):
            row_data[4] = '补办'
        else:
            row_data[4] = '首次'

        row_data[6] = '窗口'
        row_data[7] = xianxia_d if xianxia_d is not None else "无窗口办理流程"

        mid_st = '，这个'
        if xianxia_d is None:
            continue
        else:

            # 遍历要读取的列
            flag = False

            dialogue = {}

            cnt = 0

            new_ans = ['', '']

            new_ans = ['', '']
            if area in filter_data:
                new_ans.append( row_data[4])
            else:
                new_ans.append("在" + row_data[2] +  "，" + row_data[4])
            new_ans.append(xianxia_d)

            new_deal_manner = row_data[6]

            for i in range(4):
                cnt += 1
                ans = ''
                if i == 0:
                    count = random.randint(0, 4)
                    ans = template_pre[i][count] + new_deal_manner +'办' + row_data[i]
                    if count == 2:
                        ans += mid_st + row_data[i]
                    ans += template_suf[i][count]
                elif i == 1:
                    if area in filter_data:
                        ans = template_pre[i][1] + template_suf[i]
                    else:
                        ans = template_pre[i][0] + template_suf[i]
                else:
                    ans = template_pre[i] + new_ans[i] + template_suf[i]

                dialogue[columns[i % 2]] = ans
                if len(dialogue) == 2:
                    dialogue_data.append(dialogue)
                    dialogue = {}

            all_data.append(dialogue_data)

        dialogue_data = []
        cnt = 0
        flag = False

        row_data[6] = '线上'
        row_data[7] = wet_d if wet_d is not None else "无线上办理流程"
        if wet_d is None:
            continue
        new_ans = ['', '']

        if area in filter_data:
            new_ans.append(row_data[4])
        else:
            new_ans.append("在" + row_data[2] + "，" + row_data[4])
        new_ans.append(xianxia_d)

        new_deal_manner = row_data[6]
        for i in range(4):
            cnt += 1
            ans = ''
            if i == 0:
                count = random.randint(0, 4)
                ans = template_pre[i][count] + new_deal_manner + '办' + row_data[i]
                if count == 2:
                    ans += mid_st + row_data[i]
                ans += template_suf[i][count]
            elif i == 1:
                if area in filter_data:
                    ans = template_pre[i][1] + template_suf[i]
                else:
                    ans = template_pre[i][0] + template_suf[i]
            else:
                ans = template_pre[i] + new_ans[i] + template_suf[i]

            dialogue[columns[i % 2]] = ans
            if len(dialogue) == 2:
                dialogue_data.append(dialogue)
                dialogue = {}

        all_data.append(dialogue_data)


# 待实现
def get_all_question_data():
    global all_data
    # 打开Excel文件
    workbook = openpyxl.load_workbook('./data/train.xlsx')

    # 选择要读取的工作表
    worksheet = workbook['单事项信息']

    # 项目名称[7]，网上办理流程说明[35],窗口办理流程说明[36],地区[1]

    filter_data = ['广东省', '佛山市']

    columns = ['Human', 'Model']
    template_pre = [['我要办', '我要办', '我打算办', '办', '怎么办'],
                    ['请问您在那个区域，您是线上办理还是窗口办理，您是首次办理、到期换证还是补办？',
                     '请问您是线上办理还是窗口办理，您是首次办理、到期换证还是补办？'], '',
                    '根据您的需求，您可以有如下方式办理：']
    template_suf = [['，我要办理那些东西', '', '有那些问题要办', '', ''], '', '', '']

    # 正则表达

    exchange_r = r"换发"
    comple_r = r"补发"

    # 遍历每一行数据
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        dialogue_data = []
        envent = row[7]

        wet_d = row[35]

        xianxia_d = row[36]

        area = row[1]

        row_data = [''] * 10

        for i in range(2):
            row_data[i] = envent

        row_data[2] = area

        if re.search(exchange_r, envent):
            row_data[4] = '换证'
        elif re.search(comple_r, envent):
            row_data[4] = '补办'
        else:
            row_data[4] = '首次'

        row_data[6] = '窗口'
        row_data[7] = xianxia_d if xianxia_d is not None else "无窗口办理流程"

        mid_st = '，这个'
        if xianxia_d is None:
            continue
        else:

            # 遍历要读取的列
            flag = False

            dialogue = {}

            cnt = 0

            new_ans = ['', '']
            if area in filter_data:
                new_ans.append(row_data[6] + "，" + row_data[4])
            else:
                new_ans.append("在" + row_data[2] + "，" + row_data[6] + "，" + row_data[4])
            new_ans.append(xianxia_d)
            for i in range(4):
                cnt += 1
                ans = ''
                if i == 0:
                    count = random.randint(0, 4)
                    ans = template_pre[i][count] + row_data[i]
                    if count == 2:
                        ans += mid_st + row_data[i]
                    ans += template_suf[i][count]
                elif i == 1:
                    if area in filter_data:
                        ans = template_pre[i][1] + template_suf[i]
                    else:
                        ans = template_pre[i][0] + template_suf[i]
                else:
                    ans = template_pre[i] + new_ans[i] + template_suf[i]

                dialogue[columns[i % 2]] = ans
                if len(dialogue) == 2:
                    dialogue_data.append(dialogue)
                    dialogue = {}

            all_data.append(dialogue_data)

        dialogue_data = []
        cnt = 0
        flag = False

        row_data[6] = '线上'
        row_data[7] = wet_d if wet_d is not None else "无线上办理流程"
        if wet_d is None:
            continue
        new_ans = ['', '']
        if area in filter_data:
            new_ans.append(row_data[6] + "，" + row_data[4])
        else:
            new_ans.append("在" + row_data[2] + "，" + row_data[6] + "，" + row_data[4])
        new_ans.append(wet_d)
        for i in range(4):
            cnt += 1
            ans = ''
            if i == 0:
                count = random.randint(0, 4)
                ans = template_pre[i][count] + row_data[i]
                if count == 2:
                    ans += mid_st + row_data[i]
                ans += template_suf[i][count]
            elif i == 1:
                if area in filter_data:
                    ans = template_pre[i][1] + template_suf[i]
                else:
                    ans = template_pre[i][0] + template_suf[i]
            else:
                ans = template_pre[i] + new_ans[i] + template_suf[i]

            dialogue[columns[i % 2]] = ans
            if len(dialogue) == 2:
                dialogue_data.append(dialogue)
                dialogue = {}

        all_data.append(dialogue_data)


def get_merge_data():
    global all_data

    pa = './data/test'
    sup = '.json'
    arr = ['1','2','3']
    for i in range(3):
        with open(pa+arr[i]+sup,'r',encoding='utf-8')as fp:
            for line in json.load(fp) :
                all_data.append(line)

if __name__ == '__main__':


    # get_al_data()
    # get_area_data()
    # get_windowsOrOffilen_data()
    # 将对话数据保存为JSON文件
    get_merge_data()
    with open('./data/test_all.json', 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=4)